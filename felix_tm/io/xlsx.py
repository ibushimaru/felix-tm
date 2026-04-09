"""Excel (xlsx) import/export.

Supports importing translation pairs from xlsx files with configurable
source/target column mapping, and exporting TM records to xlsx.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from ..memory.record import Record


def import_xlsx(
    path: str | Path,
    source_col: int = 1,
    target_col: int = 2,
    sheet: str | int | None = None,
    header_row: int = 1,
    encoding: str | None = None,
) -> list[Record]:
    """Import records from an xlsx file.

    Args:
        path: Path to xlsx file.
        source_col: 1-based column index for source text.
        target_col: 1-based column index for target text.
        sheet: Sheet name or 0-based index. None for active sheet.
        header_row: Row number of the header (data starts at header_row + 1).
        encoding: Not used for xlsx (always UTF-8), kept for API consistency.

    Returns:
        List of Record objects.
    """
    wb = load_workbook(str(path), read_only=True, data_only=True)

    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]

    records: list[Record] = []

    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), 1):
        if len(row) < max(source_col, target_col):
            continue

        src = row[source_col - 1]
        tgt = row[target_col - 1]

        # Convert to string, skip empty
        src = str(src).strip() if src is not None else ""
        tgt = str(tgt).strip() if tgt is not None else ""

        if not src and not tgt:
            continue

        records.append(Record(source=src, target=tgt))

    wb.close()
    return records


def export_xlsx(
    records: list[Record],
    path: str | Path,
    source_lang: str = "Source",
    target_lang: str = "Target",
) -> None:
    """Export records to an xlsx file.

    Args:
        records: List of Record objects.
        path: Output file path.
        source_lang: Header label for source column.
        target_lang: Header label for target column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Translation Memory"

    # Header
    headers = [
        source_lang, target_lang, "Context", "Reliability",
        "Validated", "Ref Count", "Created", "Modified",
        "Created By", "Modified By",
    ]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data
    for rec in records:
        ws.append([
            rec.source,
            rec.target,
            rec.context,
            rec.reliability,
            "Yes" if rec.validated else "No",
            rec.refcount,
            rec.created.isoformat() if rec.created else "",
            rec.modified.isoformat() if rec.modified else "",
            rec.created_by,
            rec.modified_by,
        ])

    # Auto-width for first two columns
    for col_idx in (1, 2):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, min(ws.max_row + 1, 102))),
            default=10,
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    wb.save(str(path))
